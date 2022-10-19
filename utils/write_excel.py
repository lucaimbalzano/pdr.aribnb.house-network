# Dreams without Goals are just Dreams
#
# - @lucaimbalzano



import traceback
from importlib.resources import path
from openpyxl import Workbook, worksheet, load_workbook
from os import listdir
import os
import openpyxl
from settings.settings import PREFIX_URL_EXCEL
from console.log.logger import get_logger

logger = get_logger()

def write_excel_by_column_deprecated(start_cell_column,start_cell_row,  data_to_write, ws):
    # C14:38 
    start_cell = 'C'
    start_cell_row = 14
    for page in range(len(data_to_write)):
        for i in range(start_cell_row, page):
            if(i % 2 == 0):
                # print("[DEBUG] - price: "+[data_to_write[page][i].price])
                ws.append([data_to_write[page][i].price])    
            else:
                print("[DEBUG] - url: "+[data_to_write[page][i].url])
                ws.append(PREFIX_URL_EXCEL + [data_to_write[page][i].url])
                
                 
def write_excel_by_column(row_index, house_list_container, ws, letter):
    house_list_container_index = len(house_list_container) - 1
    # for row_index in range(row_index, len(house_list_container)*len(house_list_container[0])):
    try:
        row_index = 13
        for page in (0, house_list_container_index):
            for index_house in range(0, len(house_list_container[page])):
                house = house_list_container[page][index_house]
                row_index = row_index + 1
                if (row_index % 2 == 0):
                    logger.debug("price: "+str(house.price))
                    cell_to_write = letter + str(row_index)
                    ws[cell_to_write] = int(house.price)
                else:
                    logger.debug("url: " + str(house.url))
                    cell_to_write = letter + str(row_index)
                    ws[cell_to_write] = PREFIX_URL_EXCEL + house.url

    except Exception as e:
        print("write-excel::write_excel_by_column - "+str(e))
        logger.error("write-excel::write_excel_by_column - "+str(e))
        traceback.print_exc()
    finally:
        None

def open_excel(path):
    wb = openpyxl.load_workbook(path)
    return wb

def close_excel(path, wb):
    assert os.path.isfile(path)
    if ( wb is None):
        logger.error('Error occurred - Workbook excel is None')
        print('Error occurred - Workbook excel is None')
    wb.save(path)
    wb.close()

def write_excel_by_data_retrived(houses_airbnb, letter, wb):
    path = "C:\\Users\\lucai\\Documents\\Workspaces\\house-network\\pdr.aribnb.house-network\\media\\3_STANDARD_PianoDiRendimentoy.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb['Raccolta_Dati_Airbnb_0']


    # 1MONTH  1MONTHWEEK 3MONTH  3MONTHWEEK 6MONTH  6MONTHWEEK
    # C14:38  D14:38     E14:38   F14:38    G14:38  H14:38    [X14:76]
    write_excel_by_column( 13,houses_airbnb,ws, letter )

    wb.save(path)
    wb.close()
